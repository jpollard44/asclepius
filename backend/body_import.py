"""Parse and import smart-scale body-composition exports.

Smart scales (Renpho, Withings, and similar) export a wide table — one row per
weigh-in — with columns like Weight(kg), Body Fat, Muscle Mass, BMR and so on.
Asclepius stores metrics in a long/EAV ``daily_metrics`` table (one row per
metric per day), so this module reshapes each weigh-in into one value per
metric and writes them with source='scale'.

Both the one-off TSV import script and the ``POST /api/import/body`` xlsx
endpoint share the parsing, cleaning and de-duplication here so they behave
identically.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path

from .store import SOURCE_SCALE
from . import tracking

# Pounds -> kilograms. Several scale columns ("103.6lb") are reported in lb but
# Asclepius stores body weights in kg.
LB_TO_KG = 1.0 / 2.2046226218

# Cell sentinels the scale writes when a reading is missing.
_BLANKS = {"", "-", "--", "- -", "n/a", "na", "none"}


class BodyImportError(Exception):
    """Raised when a body-composition file can't be read or parsed."""


# ---------------------------------------------------------------------------
# Cell cleaning / value parsing
# ---------------------------------------------------------------------------
def _clean(cell) -> str | None:
    """Normalise a raw cell to a trimmed string, or None for missing data."""
    if cell is None:
        return None
    s = str(cell).strip()
    if s.lower() in _BLANKS:
        return None
    return s


def _num(cell) -> float | None:
    """Plain number (e.g. BMI, visceral fat, BMR, metabolic age)."""
    s = _clean(cell)
    if s is None:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _pct(cell) -> float | None:
    """A percentage like "15.0%" -> 15.0 (the % is stripped)."""
    s = _clean(cell)
    if s is None:
        return None
    try:
        return float(s.rstrip("%").strip())
    except ValueError:
        return None


def _kg(cell) -> float | None:
    """A weight like "59.0kg" -> 59.0 (already kilograms)."""
    s = _clean(cell)
    if s is None:
        return None
    s = s.lower().replace("kg", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _lb_to_kg(cell) -> float | None:
    """A weight like "103.6lb" -> kilograms."""
    s = _clean(cell)
    if s is None:
        return None
    s = s.lower().replace("lbs", "").replace("lb", "").strip()
    try:
        return round(float(s) * LB_TO_KG, 3)
    except ValueError:
        return None


# Source column header -> (metric key, parser). Columns not listed (Number,
# Date and Time, Weight(lb)) are ignored: weight is taken from the kg column.
COLUMN_MAP: dict[str, tuple[str, object]] = {
    "Weight(kg)": ("body_mass", _kg),
    "BMI": ("bmi", _num),
    "Body Fat": ("body_fat", _pct),
    "Muscle Mass": ("muscle_mass", _lb_to_kg),
    "Muscle Mass %": ("muscle_mass_pct", _pct),
    "Body Water": ("body_water", _pct),
    "Lean Body Mass": ("lean_body_mass", _lb_to_kg),
    "Bone Mass": ("bone_mass", _lb_to_kg),
    "Protein": ("protein_pct", _pct),
    "Visceral Fat": ("visceral_fat", _num),
    "BMR": ("bmr", _num),
    "Metabolic Age": ("metabolic_age", _num),
    "Skeletal Muscle Rate %": ("skeletal_muscle", _pct),
    "Fat Content": ("fat_content", _lb_to_kg),
    "Subcutaneous Fat": ("subcutaneous_fat", _pct),
}

_DATE_COLUMN = "Date and Time"

# Formats seen across scale exports. The primary one is "2026.05.10 08:34 PM".
_DATE_FORMATS = (
    "%Y.%m.%d %I:%M %p",
    "%Y.%m.%d %H:%M",
    "%Y-%m-%d %I:%M %p",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
)


def _parse_datetime(cell) -> datetime | None:
    if cell is None:
        return None
    if isinstance(cell, datetime):
        return cell
    s = str(cell).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Table parsing & de-duplication
# ---------------------------------------------------------------------------
def parse_table(headers: list[str], rows: list) -> list[dict]:
    """Reshape raw rows into per-weigh-in records.

    Each record is ``{date, datetime, metrics: {key: value}, n}`` where ``n`` is
    the count of non-null metrics (used to pick the richest of duplicate rows).
    Rows without a parseable date are skipped.
    """
    headers = [(_clean(h) or "") for h in headers]
    idx = {h: i for i, h in enumerate(headers)}
    if _DATE_COLUMN not in idx:
        raise BodyImportError(
            f"Missing required column '{_DATE_COLUMN}'. "
            f"Found columns: {', '.join(h for h in headers if h) or '(none)'}.")

    records: list[dict] = []
    for raw in rows:
        if raw is None:
            continue
        cells = list(raw)
        if all(_clean(c) is None for c in cells):
            continue  # blank line

        def cell(col: str):
            i = idx.get(col)
            return cells[i] if i is not None and i < len(cells) else None

        dt = _parse_datetime(cell(_DATE_COLUMN))
        if dt is None:
            continue

        metrics: dict[str, float] = {}
        for col, (key, parser) in COLUMN_MAP.items():
            if col not in idx:
                continue
            val = parser(cell(col))
            if val is not None:
                metrics[key] = val
        records.append({
            "date": dt.date().isoformat(),
            "datetime": dt.isoformat(timespec="minutes"),
            "metrics": metrics,
            "n": len(metrics),
        })
    return records


def dedupe(records: list[dict]) -> tuple[list[dict], int]:
    """Collapse to one record per date, keeping the most-complete one.

    ``daily_metrics`` keys on (metric, date), so multiple weigh-ins on the same
    calendar day would overwrite each other anyway; we deterministically keep the
    row with the most populated metrics (ties broken by the later time). Returns
    the kept records (sorted by date) and the number of rows dropped.
    """
    best: dict[str, dict] = {}
    dropped = 0
    for rec in records:
        cur = best.get(rec["date"])
        if cur is None:
            best[rec["date"]] = rec
        else:
            dropped += 1
            if (rec["n"], rec["datetime"]) > (cur["n"], cur["datetime"]):
                best[rec["date"]] = rec
    kept = sorted(best.values(), key=lambda r: r["date"])
    return kept, dropped


def import_records(records: list[dict], source: str = SOURCE_SCALE,
                   db_path: Path | None = None) -> dict:
    """De-duplicate parsed records and write them to daily_metrics.

    Returns a summary suitable for an API response or a CLI report.
    """
    kept, dropped = dedupe(records)
    metric_counts: dict[str, int] = {}
    dates: list[str] = []
    for rec in kept:
        written = tracking.log_body_metrics(rec["date"], rec["metrics"],
                                            source=source, db_path=db_path)
        if written:
            dates.append(rec["date"])
        for key in written:
            metric_counts[key] = metric_counts.get(key, 0) + 1
    dates.sort()
    return {
        "status": "ok",
        "rows_parsed": len(records),
        "duplicates_skipped": dropped,
        "dates_imported": len(dates),
        "values_written": sum(metric_counts.values()),
        "date_range": {"start": dates[0], "end": dates[-1]} if dates else None,
        "metrics": metric_counts,
    }


# ---------------------------------------------------------------------------
# Entry points
# ---------------------------------------------------------------------------
def parse_tsv(text: str) -> list[dict]:
    """Parse the tab-separated scale export used by the one-off import script."""
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        return []
    headers = lines[0].split("\t")
    rows = [ln.split("\t") for ln in lines[1:]]
    return parse_table(headers, rows)


def import_tsv(text: str, source: str = SOURCE_SCALE,
               db_path: Path | None = None) -> dict:
    return import_records(parse_tsv(text), source=source, db_path=db_path)


def import_xlsx(raw: bytes, sheet_name: str | None = None,
                source: str = SOURCE_SCALE, db_path: Path | None = None) -> dict:
    """Read an .xlsx export and import it. ``sheet_name`` defaults to the first."""
    import openpyxl  # local import: only needed for the xlsx path

    try:
        wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface a clean message to the UI
        raise BodyImportError(f"Could not open the spreadsheet: {exc}") from exc
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise BodyImportError(
                    f"No sheet named '{sheet_name}'. "
                    f"Sheets: {', '.join(wb.sheetnames)}.")
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        raise BodyImportError("The sheet is empty.")
    headers = list(rows[0])
    summary = import_records(parse_table(headers, rows[1:]),
                             source=source, db_path=db_path)
    summary["sheet"] = ws.title
    return summary
